# -*- coding: utf-8 -*-
"""
生成转辙机 8s 三相动作电流数据 → Excel (采样频率 100Hz)

形态基准取自 data/raw_blocking/raw/week_01.mat 的真实正常动作记录:
  0~0.2s  启动段     0 → ~3.1A (启动冲击)
  0.2~0.5s 解锁段     ~3.0A 短暂保持后回落
  0.5~4.6s 转换段     ~1.5~2.1A 持续, 带交流纹波
  4.7~5.3s 缓放段     快速跌至 ~0.3~0.4A 缓放台阶, 短暂保持
  5.3~5.5s 落零       跌至 0
  5.5~8s   零填充

三相 (A/B/C): 时间错相 + 幅度微差 + 纹波 120° 相移.
"""
import numpy as np
from scipy.interpolate import PchipInterpolator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

FS = 25             # 采样频率 Hz (25Hz → 8s = 200 点, 便于导入图片)
DUR = 8.0           # 时长 s
N = int(FS * DUR)   # 200 点
RNG = np.random.default_rng(42)

# ---- 正常动作包络控制点 (时间 s, 电流 A) — 相位 A 基准 ----
CTRL = np.array([
    [0.00, 0.02],
    [0.20, 3.05],   # 启动冲击峰值
    [0.45, 2.90],   # 解锁
    [0.70, 2.10],   # 转换段开始回落
    [1.10, 1.55],   # 转换段早期低谷
    [1.50, 1.90],
    [2.00, 2.05],
    [3.00, 1.90],
    [4.00, 1.90],
    [4.60, 1.95],   # 转换段结束
    [4.80, 0.42],   # 缓放台阶 (快速跌落)
    [5.05, 0.30],   # 缓放台阶保持
    [5.25, 0.10],   # 开始落零
    [5.45, 0.00],   # 断相
    [8.00, 0.00],
])

def envelope(t):
    """基准包络 (A 相), PCHIP 保单调插值, 尾部裁零."""
    ip = PchipInterpolator(CTRL[:, 0], CTRL[:, 1], extrapolate=False)
    e = ip(t)
    e = np.nan_to_num(e, nan=0.0)
    return np.clip(e, 0.0, None)

# 三相错相 (s) / 峰值微差 / 纹波相位偏移
# 25Hz 下 0.04s=1 点, 错相 1~2 点使三相在图上清晰可辨
PHASE_CFG = [
    ('A', 0.00, 1.000, 0.0),   # 名称, 时间错相, 幅度系数, 纹波相(°)
    ('B', 0.04, 1.015, 120.0),
    ('C', 0.08, 0.980, 240.0),
]

t = np.arange(N) / FS  # 0..7.99 s

# ---- 合成三相 ----
cols = {}
for name, shift, amp, rip_deg in PHASE_CFG:
    # 包络 (时间错相) × 幅度微差
    sig = envelope(np.clip(t - shift, 0, DUR)) * amp
    # 交流纹波: 低频 (~4Hz, 幅度 2%) — 25Hz 采样 Nyquist=12.5Hz, 28Hz 会混叠成虚假波形
    ripple = sig * 0.02 * np.sin(2 * np.pi * 4 * t + np.deg2rad(rip_deg))
    # 微小测量噪声
    noise = RNG.normal(0, 0.015, size=N)
    y = np.clip(sig + ripple + noise, 0.0, None)
    # 动作结束后 (包络=0) 硬置零, 与真实记录的零填充一致
    y[sig <= 0.0] = 0.0
    cols[name] = np.round(y, 3)

# ---- 写 Excel ----
wb = Workbook()

# Sheet1: 数据
ws = wb.active
ws.title = '三相电流'
head_font = Font(bold=True, color='FFFFFF')
head_fill = PatternFill('solid', fgColor='4472C4')
headers = ['序号', '时间(s)', 'A相电流(A)', 'B相电流(A)', 'C相电流(A)']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = Alignment(horizontal='center')
for i in range(N):
    ws.cell(row=i + 2, column=1, value=i + 1)
    ws.cell(row=i + 2, column=2, value=round(i / FS, 3))
    ws.cell(row=i + 2, column=3, value=cols['A'][i])
    ws.cell(row=i + 2, column=4, value=cols['B'][i])
    ws.cell(row=i + 2, column=5, value=cols['C'][i])
for c in range(1, 6):
    ws.column_dimensions[get_column_letter(c)].width = 12

# Sheet2: 说明
ws2 = wb.create_sheet('参数说明')
notes = [
    ['参数', '值', '说明'],
    ['设备', '电动转辙机动作电流', '三相 AC 整流后幅值包络 (正常动作)'],
    ['采样频率', f'{FS} Hz', f'8s → {N} 点, 便于导入图片'],
    ['时长', f'{DUR} s', f'{N} 个采样点 (每相)'],
    ['相数', 'A / B / C', '三相错相 0 / 40 / 80 ms, 幅度 ±2% 微差'],
    ['启动冲击峰值', f'{cols["A"].max():.2f} A', '约 0.2s 处'],
    ['转换段电流', '≈ 1.5~2.1 A', '约 0.5~4.6s 持续'],
    ['缓放台阶', '≈ 0.3~0.4 A', '约 4.8~5.1s, 快速跌落后的短暂保持'],
    ['动作总时长', '≈ 5.4 s', '之后至 8s 为零填充'],
    ['噪声/纹波', '纹波 2% @4Hz (防混叠), 高斯噪声 σ=0.015A', '25Hz 下 28Hz 会混叠, 已换低频'],
    ['随机种子', '42', '可复现'],
]
for r, row in enumerate(notes, 1):
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        if r == 1:
            cell.font = head_font
            cell.fill = head_fill
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 44

out = '转辙机三相动作电流_8s_25Hz.xlsx'
wb.save(out)

# ---- 验证波形图 ----
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as _fm
_avail = {f.name for f in _fm.fontManager.ttflist}
_cn = next((f for f in ['Microsoft YaHei', 'SimHei', 'Noto Sans SC'] if f in _avail), 'sans-serif')
plt.rcParams['font.sans-serif'] = [_cn, 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
plt.figure(figsize=(11, 3.5))
for name, *_ in PHASE_CFG:
    plt.plot(t, cols[name], lw=0.8, label=f'{name}相')
plt.xlabel('时间 (s)'); plt.ylabel('电流 (A)')
plt.title('转辙机 8s 三相动作电流 (25Hz, 正常动作)')
plt.legend(); plt.grid(alpha=0.3)
plt.tight_layout()
plt.savefig('转辙机三相动作电流_波形.png', dpi=120)
plt.close()

# ---- 打印统计 ----
print(f'已生成: {out}  ({N} 行 × 3 相)')
print(f'  A相: max={cols["A"].max():.3f}A @ {t[np.argmax(cols["A"])]:.2f}s'
      f' | 均值={cols["A"].mean():.3f}A | 非零点={(np.abs(cols["A"])>1e-3).sum()}')
print(f'  B相: max={cols["B"].max():.3f}A @ {t[np.argmax(cols["B"])]:.2f}s'
      f' | 均值={cols["B"].mean():.3f}A | 非零点={(np.abs(cols["B"])>1e-3).sum()}')
print(f'  C相: max={cols["C"].max():.3f}A @ {t[np.argmax(cols["C"])]:.2f}s'
      f' | 均值={cols["C"].mean():.3f}A | 非零点={(np.abs(cols["C"])>1e-3).sum()}')
