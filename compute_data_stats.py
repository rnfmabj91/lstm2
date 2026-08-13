# -*- coding: utf-8 -*-
"""
计算数据统计并输出 outputs/_data_stats.json (供 docx 生成脚本使用).
仅依赖 numpy / scipy / openpyxl.
"""
import json
import os
import sys
import numpy as np
from scipy import io as sio

DATA_DIR = 'data/raw_blocking/preprocessed'


def load_mat(path, key):
    return sio.loadmat(os.path.join(DATA_DIR, path))[key].astype(np.float32)


def main():
    stats = {}

    # ---- 各集样本数与测试故障数 ----
    for n in ['X_train', 'X_val', 'X_test']:
        stats[n] = load_mat(f'{n}.mat', n)
        stats[f'{n}_shape'] = list(stats[n].shape)
    labels = load_mat('labels_test.mat', 'labels_test').ravel().astype(np.int64)
    stats['n_fault'] = int(labels.sum())
    stats['n_normal_test'] = int((labels == 0).sum())

    # ---- 三相归一化电流统计 (各集抽样 2000 样本) ----
    rng = np.random.RandomState(42)
    for n in ['X_train', 'X_val', 'X_test']:
        X = stats[n]
        idx = rng.choice(len(X), min(2000, len(X)), replace=False)
        sub = X[idx]
        per_phase = []
        for c in range(3):
            vals = sub[:, c, :]
            per_phase.append({
                'mean': float(vals.mean()),
                'std': float(vals.std()),
                'p50': float(np.median(vals)),
                'p99': float(np.percentile(vals, 99)),
                'max': float(vals.max()),
            })
        stats[f'{n}_phase'] = per_phase

    # ---- 有效长度 (active region) 分布, 抽样 2000 ----
    for n in ['X_train', 'X_val', 'X_test']:
        X = stats[n]
        idx = rng.choice(len(X), min(2000, len(X)), replace=False)
        sub = X[idx]
        active = (np.abs(sub).max(axis=1) > 1e-3)          # (B, T)
        al = active.sum(axis=1)                             # (B,)
        stats[f'{n}_active_len'] = {
            'mean_pts': float(al.mean()),
            'mean_sec': float(al.mean() / 100.0),
            'p5_sec': float(np.percentile(al, 5) / 100.0),
            'p95_sec': float(np.percentile(al, 95) / 100.0),
            'min_sec': float(al.min() / 100.0),
        }
    stats['fs'] = 100
    stats['total_pts'] = 800

    # ---- 25Hz 演示 Excel 三相统计 (读生成的 xlsx) ----
    try:
        from openpyxl import load_workbook
        wb = load_workbook('转辙机三相动作电流_8s_25Hz.xlsx', read_only=True)
        ws = wb['三相电流']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        cols = {k: [] for k in ['A', 'B', 'C']}
        for r in rows:
            for i, k in enumerate(['A', 'B', 'C']):
                cols[k].append(float(r[2 + i]))
        excel_phase = {}
        for k in ['A', 'B', 'C']:
            v = np.array(cols[k])
            excel_phase[k] = {
                'max': float(v.max()),
                'mean': float(v[v > 1e-3].mean()) if (v > 1e-3).any() else 0.0,
                'nonzero': int((np.abs(v) > 1e-3).sum()),
            }
        stats['excel_25hz'] = {
            'fs': 25,
            'n_points': len(rows),
            'duration_s': 8.0,
            'phase': excel_phase,
            'faults': '正常动作',
        }
    except Exception as e:
        stats['excel_25hz'] = {'error': str(e)}

    # 释放大数组
    for n in ['X_train', 'X_val', 'X_test']:
        del stats[n]

    os.makedirs('outputs', exist_ok=True)
    with open('outputs/_data_stats.json', 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    print('[ok] outputs/_data_stats.json 已写入')
    print(json.dumps(stats, ensure_ascii=False, indent=2)[:2000])


if __name__ == '__main__':
    main()
